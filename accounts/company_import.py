"""Company data import from a 5-sheet Excel workbook, used at registration.

Sheets: Items & Stock, Vendors & Prices, Customers, Recipes (BOM),
Production Lines. Everything is scoped to the importing user's company.
The importer validates the whole workbook first and only commits if there
are zero errors — otherwise the caller re-uploads.
"""
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction

# --- Sheet + column definitions (single source of truth) -------------------

SHEETS = {
    "Items & Stock": [
        "item_code", "item_name", "category", "unit", "selling_price",
        "warehouse", "opening_quantity",
    ],
    "Vendors & Prices": [
        "vendor_name", "category", "email", "phone", "address", "rating",
        "supplies_item_code", "item_name", "unit_price", "currency",
        "min_order_qty", "lead_time_days",
    ],
    "Customers": ["customer_name", "email", "phone", "address"],
    "Recipes (BOM)": [
        "product_code", "product_name", "ingredient_code", "ingredient_name",
        "quantity_per_unit",
    ],
    "Production Lines": ["line_name", "location", "capacity_per_hour"],
}

# Read-only helper columns the importer ignores (auto VLOOKUP name columns)
IGNORED_COLUMNS = {"item_name_lookup", "product_name", "ingredient_name"}

EXAMPLE_ROWS = {
    "Items & Stock": [
        ["RM-WIRE-20", "Spring Steel Wire 2.0mm", "raw_material", "kg", "", "Raw Material Store", 3500],
        ["FG-SPR-HD40", "Compression Spring HD-40", "finished_good", "unit", 22, "Finished Goods Store", 3150],
    ],
    "Vendors & Prices": [
        ["Tata Steel Wiron", "raw_material", "sales@tatawiron.com", "+91-657-2345678",
         "Jamshedpur", 4.8, "RM-WIRE-20", "Spring Steel Wire 2.0mm", 96, "USD", 500, 7],
    ],
    "Customers": [
        ["Godrej Appliances", "purchase@godrej.com", "+91-22-67961700", "Mumbai"],
    ],
    "Recipes (BOM)": [
        ["FG-SPR-HD40", "Compression Spring HD-40", "RM-WIRE-20", "Spring Steel Wire 2.0mm", 0.12],
    ],
    "Production Lines": [
        ["Spring Coiling Line 1", "Plant 1 - Bay A", 450],
    ],
}

CATEGORY_CHOICES = ["raw_material", "finished_good"]


# --- Template generation ---------------------------------------------------

def build_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    example_font = Font(italic=True, color="808080")

    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for c, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(c)].width = max(14, len(col) + 3)
        for r, row in enumerate(EXAMPLE_ROWS.get(sheet_name, []), start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = example_font
        ws.freeze_panes = "A2"

    # Category dropdown on Items & Stock
    items_ws = wb["Items & Stock"]
    cat_col = SHEETS["Items & Stock"].index("category") + 1
    dv = DataValidation(type="list", formula1='"raw_material,finished_good"', allow_blank=False)
    items_ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(cat_col)}2:{get_column_letter(cat_col)}1000")

    # Item-code dropdowns on Vendors & Prices and Recipes, sourced from Sheet 1
    code_col = SHEETS["Items & Stock"].index("item_code") + 1
    code_letter = get_column_letter(code_col)
    code_range = f"'Items & Stock'!${code_letter}$2:${code_letter}$1000"

    def add_code_dropdown(ws, col_name):
        col = SHEETS[ws.title].index(col_name) + 1
        dv = DataValidation(type="list", formula1=f"={code_range}", allow_blank=True)
        ws.add_data_validation(dv)
        letter = get_column_letter(col)
        dv.add(f"{letter}2:{letter}1000")

    add_code_dropdown(wb["Vendors & Prices"], "supplies_item_code")
    add_code_dropdown(wb["Recipes (BOM)"], "product_code")
    add_code_dropdown(wb["Recipes (BOM)"], "ingredient_code")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Parsing + validation helpers -----------------------------------------

def _rows(ws, columns):
    """Yield (excel_row_number, {col: value}) for non-empty data rows."""
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {name: header.index(name) for name in columns if name in header}
    for r in range(2, ws.max_row + 1):
        values = [c.value for c in ws[r]]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {name: (values[i] if i < len(values) else None) for name, i in idx.items()}
        yield r, row


def _s(v):
    return "" if v is None else str(v).strip()


def _dec(v, field, sheet, rownum, errors, required=False, default="0"):
    if v is None or str(v).strip() == "":
        if required:
            errors.append(f"{sheet} row {rownum}: '{field}' is required.")
        return Decimal(default)
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        errors.append(f"{sheet} row {rownum}: '{field}' must be a number (got '{v}').")
        return Decimal(default)


# --- The importer ----------------------------------------------------------

def import_workbook(file_obj, company, user):
    """Validate and (if clean) import the workbook for `company`.

    Returns (ok: bool, result: dict). On errors, result={"errors": [...]}.
    On success, result={"summary": {...}} and nothing is written on failure.
    """
    from openpyxl import load_workbook
    from inventory.models import Item, Warehouse, Stock
    from procurement.models import Vendor, VendorPriceList
    from sales.models import Customer
    from production.models import ProductionLine, Recipe, RecipeIngredient

    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:
        return False, {"errors": [f"Could not read the Excel file: {exc}"]}

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        return False, {"errors": [f"Missing sheet(s): {', '.join(missing)}. "
                                  f"Please use the provided template."]}

    errors = []

    # ---- Pass 1: Items & Stock -> item defs + stock rows ----
    item_defs = {}     # code -> dict(name, category, unit, selling_price)
    stock_rows = []    # (code, warehouse_name, qty)
    seen_stock = set()
    ws = wb["Items & Stock"]
    for rn, row in _rows(ws, SHEETS["Items & Stock"]):
        code = _s(row.get("item_code"))
        if not code:
            errors.append(f"Items & Stock row {rn}: 'item_code' is required.")
            continue
        name = _s(row.get("item_name"))
        category = _s(row.get("category")).lower()
        unit = _s(row.get("unit")) or "unit"
        wh = _s(row.get("warehouse"))
        if not name:
            errors.append(f"Items & Stock row {rn}: 'item_name' is required.")
        if category not in CATEGORY_CHOICES:
            errors.append(f"Items & Stock row {rn}: 'category' must be raw_material or finished_good.")
        if not wh:
            errors.append(f"Items & Stock row {rn}: 'warehouse' is required.")
        selling = _dec(row.get("selling_price"), "selling_price", "Items & Stock", rn, errors)
        qty = _dec(row.get("opening_quantity"), "opening_quantity", "Items & Stock", rn, errors)

        if code in item_defs:
            prev = item_defs[code]
            if (name and name != prev["name"]) or (category and category != prev["category"]) or (unit != prev["unit"]):
                errors.append(f"Items & Stock row {rn}: item_code '{code}' has conflicting "
                              f"name/category/unit vs an earlier row.")
        else:
            item_defs[code] = {"name": name, "category": category, "unit": unit, "selling_price": selling}

        key = (code, wh.lower())
        if key in seen_stock:
            errors.append(f"Items & Stock row {rn}: item_code '{code}' + warehouse '{wh}' "
                          f"appears more than once.")
        seen_stock.add(key)
        if wh:
            stock_rows.append((code, wh, qty))

    # ---- Pass 2: Vendors & Prices ----
    vendor_defs = {}   # name -> dict(category,email,phone,address,rating)
    price_rows = []    # (vendor_name, item_code, unit_price, currency, moq, lead)
    ws = wb["Vendors & Prices"]
    for rn, row in _rows(ws, SHEETS["Vendors & Prices"]):
        vname = _s(row.get("vendor_name"))
        if not vname:
            errors.append(f"Vendors & Prices row {rn}: 'vendor_name' is required.")
            continue
        vendor_defs.setdefault(vname, {
            "category": _s(row.get("category")) or "raw_material",
            "email": _s(row.get("email")),
            "phone": _s(row.get("phone")),
            "address": _s(row.get("address")),
            "rating": float(_dec(row.get("rating"), "rating", "Vendors & Prices", rn, errors)),
        })
        code = _s(row.get("supplies_item_code"))
        if code:
            if code not in item_defs:
                errors.append(f"Vendors & Prices row {rn}: supplies_item_code '{code}' "
                              f"not found in Items & Stock.")
            else:
                price_rows.append((
                    vname, code,
                    _dec(row.get("unit_price"), "unit_price", "Vendors & Prices", rn, errors, required=True),
                    _s(row.get("currency")) or "USD",
                    float(_dec(row.get("min_order_qty"), "min_order_qty", "Vendors & Prices", rn, errors, default="1")),
                    int(_dec(row.get("lead_time_days"), "lead_time_days", "Vendors & Prices", rn, errors, default="7")),
                ))

    # ---- Pass 3: Customers ----
    customer_rows = []
    ws = wb["Customers"]
    for rn, row in _rows(ws, SHEETS["Customers"]):
        cname = _s(row.get("customer_name"))
        if not cname:
            errors.append(f"Customers row {rn}: 'customer_name' is required.")
            continue
        customer_rows.append({
            "name": cname, "email": _s(row.get("email")),
            "phone": _s(row.get("phone")), "address": _s(row.get("address")),
        })

    # ---- Pass 4: Production Lines ----
    line_rows = []
    ws = wb["Production Lines"]
    for rn, row in _rows(ws, SHEETS["Production Lines"]):
        lname = _s(row.get("line_name"))
        if not lname:
            errors.append(f"Production Lines row {rn}: 'line_name' is required.")
            continue
        line_rows.append({
            "name": lname, "location": _s(row.get("location")) or "Main Facility",
            "capacity": float(_dec(row.get("capacity_per_hour"), "capacity_per_hour",
                                   "Production Lines", rn, errors, default="100")),
        })

    # ---- Pass 5: Recipes (BOM) ----
    recipe_rows = []   # (product_code, ingredient_code, qty)
    ws = wb["Recipes (BOM)"]
    for rn, row in _rows(ws, SHEETS["Recipes (BOM)"]):
        pcode = _s(row.get("product_code"))
        icode = _s(row.get("ingredient_code"))
        if not pcode or not icode:
            errors.append(f"Recipes (BOM) row {rn}: 'product_code' and 'ingredient_code' are required.")
            continue
        if pcode not in item_defs:
            errors.append(f"Recipes (BOM) row {rn}: product_code '{pcode}' not in Items & Stock.")
        elif item_defs[pcode]["category"] != "finished_good":
            errors.append(f"Recipes (BOM) row {rn}: product '{pcode}' must be a finished_good.")
        if icode not in item_defs:
            errors.append(f"Recipes (BOM) row {rn}: ingredient_code '{icode}' not in Items & Stock.")
        qty = _dec(row.get("quantity_per_unit"), "quantity_per_unit", "Recipes (BOM)", rn, errors, required=True)
        if pcode in item_defs and icode in item_defs:
            recipe_rows.append((pcode, icode, qty))

    if errors:
        return False, {"errors": errors}

    # ---- Commit (all-or-nothing) ----
    from .company_import_commit import commit
    summary = commit(company, user, item_defs, stock_rows, vendor_defs,
                     price_rows, customer_rows, line_rows, recipe_rows)
    return True, {"summary": summary}
